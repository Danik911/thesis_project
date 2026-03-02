"""Excel export for BI filtered datasets."""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from .filter_engine import get_filter_engine
from .session_store import get_session


def export_excel(session_id: str) -> BytesIO:
    """Export filtered BI data to an Excel workbook."""
    session = get_session(session_id)
    engine = get_filter_engine(session_id)
    filtered_df = engine.get_filtered_dataframe()

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Data"

    columns = [str(column) for column in filtered_df.columns]

    for col_idx, column_name in enumerate(columns, start=1):
        cell = data_sheet.cell(row=1, column=col_idx, value=column_name)
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(filtered_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            data_sheet.cell(row=row_idx, column=col_idx, value=value)

    active_filters = engine.get_active_filters()
    if active_filters:
        filters_sheet = workbook.create_sheet("Filters Applied")
        filters_sheet.append(["Column", "Operator", "Value", "Applied At (UTC)"])
        for cell in filters_sheet[1]:
            cell.font = Font(bold=True)

        timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        for filter_item in active_filters:
            filters_sheet.append(
                [
                    str(filter_item.get("column", "")),
                    str(filter_item.get("operator", "")),
                    str(filter_item.get("value", "")),
                    timestamp,
                ]
            )

        filters_sheet.append([])
        filters_sheet.append(["Source File", session.filename])
        filters_sheet.append(["Total Source Rows", session.total_rows])
        filters_sheet.append(["Filtered Rows", len(filtered_df)])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
