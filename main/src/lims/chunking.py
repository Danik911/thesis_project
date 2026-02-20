"""Sheet-level chunking for LabWare XLSX MDA templates.

Replaces the monolithic ``parse_xlsx_to_text()`` approach (1 blob per XLSX)
with per-sheet markdown table chunks (~5 per XLSX, ~125 total across 25 files).

Each chunk contains:
- A single sheet rendered as a markdown table
- Metadata: source file, sheet name, priority flag, chunk ID
- One additional SUMMARY chunk per workbook with key details

Priority sheets (Analysis, Component, Calc Variable, Calculation) are flagged
so the retriever can boost them during ranking.

NO FALLBACK LOGIC.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

# Sheets most relevant for MDA generation
PRIORITY_SHEETS: set[str] = {
    "analysis",
    "component",
    "calc variable",
    "calculation",
}


def sheet_to_markdown_table(
    sheet_name: str,
    headers: list[str],
    rows: list[list[str]],
) -> str:
    """Convert a sheet's headers + rows into a markdown table string.

    Args:
        sheet_name: Name of the worksheet.
        headers: Column header strings.
        rows: List of row value lists.

    Returns:
        Markdown-formatted table with a ``## Sheet: ...`` heading.
    """
    lines: list[str] = [f"## Sheet: {sheet_name}"]

    if not headers:
        return f"## Sheet: {sheet_name}\n\n(empty)"

    # Header row
    lines.append("| " + " | ".join(headers) + " |")
    # Separator row
    lines.append("| " + " | ".join("---" for _ in headers) + " |")
    # Data rows
    for row in rows:
        # Pad row to header length, escape pipes in cell values
        padded = row + [""] * (len(headers) - len(row))
        cells = [str(c).replace("|", "\\|") for c in padded[:len(headers)]]
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines)


def _extract_sample_values(
    headers: list[str],
    rows: list[list[str]],
    max_samples: int = 5,
) -> dict[str, list[str]]:
    """Extract sample values from key columns for the summary chunk.

    Picks columns whose headers contain keywords like 'name', 'analysis',
    'component', 'type', 'method' and returns up to *max_samples* unique
    non-empty values per column.
    """
    keywords = {"name", "analysis", "component", "type", "method", "code"}
    samples: dict[str, list[str]] = {}

    for idx, header in enumerate(headers):
        header_lower = header.lower()
        if any(kw in header_lower for kw in keywords):
            values: list[str] = []
            for row in rows:
                if idx < len(row):
                    val = str(row[idx]).strip()
                    if val and val.lower() not in {"none", ""}:
                        if val not in values:
                            values.append(val)
                        if len(values) >= max_samples:
                            break
            if values:
                samples[header] = values

    return samples


def generate_workbook_summary(
    xlsx_name: str,
    sheet_summaries: list[dict[str, object]],
) -> str:
    """Build a natural-language summary chunk for an entire workbook.

    Args:
        xlsx_name: Filename of the XLSX.
        sheet_summaries: List of dicts with 'sheet_name', 'row_count',
            'headers', 'sample_values'.

    Returns:
        Summary text suitable for embedding.
    """
    lines: list[str] = [
        f"# MDA Template Summary: {xlsx_name}",
        "",
    ]

    # Derive site prefix
    prefix = xlsx_name.split("_")[0] if "_" in xlsx_name else "UNKNOWN"
    lines.append(f"Site prefix: {prefix}")
    lines.append(f"Sheets: {len(sheet_summaries)}")
    lines.append("")

    for ss in sheet_summaries:
        sheet_name = ss["sheet_name"]
        row_count = ss["row_count"]
        lines.append(f"### {sheet_name} ({row_count} rows)")

        sample_values = ss.get("sample_values", {})
        if sample_values:
            for col, vals in sample_values.items():
                lines.append(f"  - {col}: {', '.join(str(v) for v in vals)}")
        lines.append("")

    return "\n".join(lines)


def parse_xlsx_to_chunks(xlsx_path: Path | str) -> list[dict]:
    """Parse a LabWare XLSX workbook into per-sheet markdown chunks.

    Args:
        xlsx_path: Path to an ``.xlsx`` file.

    Returns:
        List of chunk dicts, each with keys:
        - ``id``: Unique chunk identifier (``{stem}__{sheet_name}`` or ``{stem}__SUMMARY``)
        - ``text``: Chunk content (markdown table or summary)
        - ``metadata``: Dict with ``source_file``, ``sheet_name``, ``is_priority``,
          ``is_summary``, ``row_count``, ``prefix``

    Raises:
        FileNotFoundError: If *xlsx_path* does not exist.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX file does not exist: {xlsx_path.resolve()}")

    logger.info("Chunking XLSX: %s", xlsx_path.name)
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    stem = xlsx_path.stem
    prefix = stem.split("_")[0] if "_" in stem else "UNKNOWN"
    chunks: list[dict] = []
    sheet_summaries: list[dict[str, object]] = []

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_raw = list(ws.iter_rows(values_only=True))

            if not rows_raw:
                continue

            # First row = headers
            headers = [str(c) if c is not None else "" for c in rows_raw[0]]
            data_rows = [
                [str(c) if c is not None else "" for c in row]
                for row in rows_raw[1:]
            ]

            # Skip sheets that are entirely empty (headers only, no data)
            if not data_rows:
                continue

            # Normalize sheet name for ID (lowercase, replace spaces)
            safe_sheet = re.sub(r"\s+", "_", sheet_name.strip().lower())
            chunk_id = f"{stem}__{safe_sheet}"

            is_priority = sheet_name.strip().lower() in PRIORITY_SHEETS

            md_table = sheet_to_markdown_table(sheet_name, headers, data_rows)

            chunks.append({
                "id": chunk_id,
                "text": md_table,
                "metadata": {
                    "source_file": xlsx_path.name,
                    "sheet_name": sheet_name,
                    "is_priority": is_priority,
                    "is_summary": False,
                    "row_count": len(data_rows),
                    "prefix": prefix,
                },
            })

            # Collect info for summary
            sample_values = _extract_sample_values(headers, data_rows)
            sheet_summaries.append({
                "sheet_name": sheet_name,
                "row_count": len(data_rows),
                "headers": headers,
                "sample_values": sample_values,
            })
    finally:
        wb.close()

    # Add workbook summary chunk
    if sheet_summaries:
        summary_text = generate_workbook_summary(xlsx_path.name, sheet_summaries)
        chunks.append({
            "id": f"{stem}__SUMMARY",
            "text": summary_text,
            "metadata": {
                "source_file": xlsx_path.name,
                "sheet_name": "SUMMARY",
                "is_priority": False,
                "is_summary": True,
                "row_count": 0,
                "prefix": prefix,
            },
        })

    logger.info(
        "Chunked %s: %d chunks (%d sheets + 1 summary)",
        xlsx_path.name,
        len(chunks),
        len(chunks) - 1 if chunks else 0,
    )

    return chunks
