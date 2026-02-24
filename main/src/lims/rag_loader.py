"""RAG loader for MDA templates -- parse XLSX files into ChromaDB.

Parses demo LabWare XLSX files into text representations and stores them
in the ChromaDB ``mda_templates`` collection for RAG-based retrieval
during MDA generation.

ChromaDB storage is kept separate from the thesis project:
  - Thesis:   ``./chroma_db/``
  - AI4LIMS:  ``./chroma_db_lims/``

NO FALLBACK LOGIC: all errors raise with full diagnostics.
"""

from __future__ import annotations

import logging
import os
import re
from pathlib import Path
from typing import Any

import chromadb
import openpyxl
from langfuse import observe
from rank_bm25 import BM25Okapi

from main.src.lims.chunking import parse_xlsx_to_chunks

logger = logging.getLogger(__name__)

CHROMA_PATH = os.getenv("LIMS_CHROMADB_PATH", "./chroma_db_lims")
COLLECTION_NAME = "mda_templates"


# ---------------------------------------------------------------------------
# 1. XLSX parsing
# ---------------------------------------------------------------------------


def parse_xlsx_to_text(xlsx_path: Path) -> str:
    """Parse a LabWare XLSX workbook into a single text string for embedding.

    Reads every sheet in the workbook, extracts header rows and data rows,
    and returns a human-readable text representation suitable for ChromaDB
    embedding and similarity search.

    Args:
        xlsx_path: Absolute or relative path to an ``.xlsx`` file.

    Returns:
        A multi-line string containing sheet names, headers, and row data.

    Raises:
        FileNotFoundError: If *xlsx_path* does not exist.
        openpyxl.utils.exceptions.InvalidFileException: If the file is not
            a valid XLSX workbook.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        msg = f"XLSX file does not exist: {xlsx_path.resolve()}"
        raise FileNotFoundError(msg)

    logger.info("Parsing XLSX: %s", xlsx_path.name)
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    parts: list[str] = [f"File: {xlsx_path.name}"]

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                parts.append(f"\nSheet: {sheet_name}\n(empty)")
                continue

            # First row is treated as the header
            headers = [str(cell) if cell is not None else "" for cell in rows[0]]
            parts.append(f"\nSheet: {sheet_name}")
            parts.append(f"Headers: {' | '.join(headers)}")

            # Data rows
            for row_idx, row in enumerate(rows[1:], start=1):
                values = [
                    str(cell) if cell is not None else "" for cell in row
                ]
                parts.append(f"Row {row_idx}: {' | '.join(values)}")
    finally:
        wb.close()

    text = "\n".join(parts)
    logger.info(
        "Parsed %s: %d characters across %d sheets",
        xlsx_path.name,
        len(text),
        len(wb.sheetnames),
    )
    return text


# ---------------------------------------------------------------------------
# 2. Seed ChromaDB from demo XLSX files
# ---------------------------------------------------------------------------


def seed_mda_templates(
    demo_data_dir: str = "./demo_data",
    collection_name: str = COLLECTION_NAME,
    chroma_path: str = CHROMA_PATH,
) -> int:
    """Seed a ChromaDB collection from demo XLSX files using sheet-level chunks.

    Iterates over all ``*.xlsx`` files in *demo_data_dir*, chunks each into
    per-sheet markdown tables via :func:`parse_xlsx_to_chunks`, and upserts
    the resulting chunks into the target collection.

    Args:
        demo_data_dir: Directory containing demo ``*.xlsx`` files.
        collection_name: ChromaDB collection name (defaults to ``mda_templates``).
        chroma_path: Path to the ChromaDB persistent storage directory.

    Returns:
        Number of chunks added to the collection.

    Raises:
        FileNotFoundError: If *demo_data_dir* does not exist or contains
            no ``.xlsx`` files.
    """
    demo_dir = Path(demo_data_dir)
    if not demo_dir.exists():
        msg = f"Demo data directory does not exist: {demo_dir.resolve()}"
        raise FileNotFoundError(msg)

    xlsx_files = sorted(demo_dir.glob("*.xlsx"))
    if not xlsx_files:
        msg = (
            f"No .xlsx files found in {demo_dir.resolve()}. "
            f"Expected LabWare demo XLSX files with AND_*, FRE_*, TUA_* prefixes."
        )
        raise FileNotFoundError(msg)

    logger.info(
        "Found %d XLSX files in %s", len(xlsx_files), demo_dir.resolve()
    )

    client = chromadb.PersistentClient(path=chroma_path)
    collection = client.get_or_create_collection(collection_name)

    documents: list[str] = []
    metadatas: list[dict[str, Any]] = []
    ids: list[str] = []

    for xlsx_path in xlsx_files:
        chunks = parse_xlsx_to_chunks(xlsx_path)
        for chunk in chunks:
            documents.append(chunk["text"])
            # ChromaDB metadata values must be str, int, float, or bool
            meta = {
                "source_file": chunk["metadata"]["source_file"],
                "sheet_name": chunk["metadata"]["sheet_name"],
                "is_priority": chunk["metadata"]["is_priority"],
                "is_summary": chunk["metadata"]["is_summary"],
                "row_count": chunk["metadata"]["row_count"],
                "prefix": chunk["metadata"]["prefix"],
            }
            metadatas.append(meta)
            ids.append(chunk["id"])

    # Bulk upsert to ChromaDB (idempotent -- safe to re-run)
    collection.upsert(documents=documents, metadatas=metadatas, ids=ids)

    final_count = collection.count()
    logger.info(
        "Seeded %d chunks from %d XLSX files into '%s' (total: %d)",
        len(documents),
        len(xlsx_files),
        collection_name,
        final_count,
    )

    return len(documents)


# ---------------------------------------------------------------------------
# 3. Query similar templates
# ---------------------------------------------------------------------------

_bm25_index: BM25Okapi | None = None
_bm25_documents: list[str] = []
_bm25_doc_ids: list[str] = []

_RAG_QUERY_SYNONYMS: dict[str, list[str]] = {
    "identity": ["id", "identification"],
    "control": ["ctl"],
    "meta": ["metadata"],
    "calculation": ["calc"],
    "component": ["components"],
    "suitability": ["sfu", "suitability_for_use"],
}

_SHEET_INTENT_TERMS: dict[str, set[str]] = {
    "analysis": {"analysis", "identity", "id", "method"},
    "component": {"component", "components", "control", "ctl", "meta"},
    "calc variable": {"variable", "variables", "calc", "calculation"},
    "calculation": {"calc", "calculation", "formula", "equation"},
    "summary": {"overview", "summary"},
}


def _build_bm25_index(collection: chromadb.Collection) -> None:
    """Build BM25 index from all documents in the ChromaDB collection."""
    global _bm25_index, _bm25_documents, _bm25_doc_ids
    if _bm25_index is not None:
        return

    results = _collection_get(collection, include=["documents"])

    _bm25_documents = results.get("documents", [])
    _bm25_doc_ids = results.get("ids", [])

    tokenized_corpus = [_tokenize(doc) for doc in _bm25_documents]
    _bm25_index = BM25Okapi(tokenized_corpus)
    logger.info("BM25 index built with %d documents", len(_bm25_documents))


def _tokenize(text: str) -> list[str]:
    """Simple tokenizer: lowercase, split on non-alphanumeric."""
    return re.findall(r"\w+", text.lower())


def _reciprocal_rank_fusion(
    semantic_ids: list[str],
    bm25_ids: list[str],
    semantic_weight: float = 0.6,
    bm25_weight: float = 0.4,
    k: int = 60,
) -> list[str]:
    """Combine semantic and BM25 rankings using Reciprocal Rank Fusion."""
    scores: dict[str, float] = {}

    for rank, doc_id in enumerate(semantic_ids):
        scores[doc_id] = scores.get(doc_id, 0.0) + semantic_weight * (1.0 / (k + rank))

    for rank, doc_id in enumerate(bm25_ids):
        scores[doc_id] = scores.get(doc_id, 0.0) + bm25_weight * (1.0 / (k + rank))

    sorted_ids = sorted(scores.keys(), key=lambda x: scores[x], reverse=True)
    return sorted_ids


def _extract_site_prefix(text: str) -> str | None:
    """Extract site prefix from extraction text for pre-filtering."""
    match = re.search(r"\b(AND|FRE|TUA)\b", text, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return None


def _extract_method_identifiers(text: str) -> list[str]:
    """Extract likely method identifiers such as AND_ACS_DYE from query text."""
    matches = re.findall(r"\b[A-Z]{3}_[A-Z0-9]+(?:_[A-Z0-9]+){1,3}\b", text.upper())
    deduped: list[str] = []
    seen: set[str] = set()
    for item in matches:
        if item not in seen:
            seen.add(item)
            deduped.append(item)
    return deduped


def _collection_get(
    collection: chromadb.Collection,
    *,
    include: list[str],
    ids: list[str] | None = None,
) -> dict[str, Any]:
    """Retrieve collection payload with compatibility for test fakes."""
    if hasattr(collection, "get"):
        return collection.get(ids=ids, include=include)

    if not (hasattr(collection, "_documents") and hasattr(collection, "_ids")):
        raise AttributeError(
            "Collection object does not support 'get()' and does not expose "
            "test fixture fields _documents/_ids."
        )

    doc_ids = list(collection._ids)
    docs = list(collection._documents)
    metas = list(getattr(collection, "_metadatas", [{} for _ in doc_ids]))

    index_by_id = {doc_id: idx for idx, doc_id in enumerate(doc_ids)}
    selected_ids = ids if ids is not None else doc_ids

    out_ids: list[str] = []
    out_docs: list[str] = []
    out_metas: list[dict[str, Any]] = []
    for doc_id in selected_ids:
        idx = index_by_id.get(doc_id)
        if idx is None:
            continue
        out_ids.append(doc_id)
        out_docs.append(docs[idx])
        out_metas.append(metas[idx] if idx < len(metas) else {})

    payload: dict[str, Any] = {"ids": out_ids}
    if "documents" in include:
        payload["documents"] = out_docs
    if "metadatas" in include:
        payload["metadatas"] = out_metas
    return payload


def _build_augmented_queries(
    query: str,
    max_queries: int,
) -> list[str]:
    """Build deterministic query variants for higher recall with bounded fan-out."""
    normalized = " ".join(query.split())
    variants: list[str] = [normalized]

    lower = normalized.lower()
    expanded_terms: list[str] = []
    for term, synonyms in _RAG_QUERY_SYNONYMS.items():
        if re.search(rf"\b{re.escape(term)}\b", lower):
            expanded_terms.extend(synonyms)

    if expanded_terms:
        variants.append(f"{normalized} {' '.join(expanded_terms)}")

    method_ids = _extract_method_identifiers(normalized)
    for method_id in method_ids:
        variants.append(f"{method_id} component calculation")
        variants.append(f"{method_id} analysis summary")

    if re.search(r"\b(control|ctl|meta|identity)\b", lower):
        variants.append(f"{normalized} component analysis")

    deduped: list[str] = []
    seen: set[str] = set()
    for variant in variants:
        compact = " ".join(variant.split())
        key = compact.lower()
        if compact and key not in seen:
            seen.add(key)
            deduped.append(compact)
        if len(deduped) >= max_queries:
            break

    return deduped


def _tokens_for_matching(text: str) -> set[str]:
    """Extract uppercase tokens useful for metadata match boosting."""
    tokens = {t.upper() for t in re.findall(r"[A-Za-z0-9_]+", text) if len(t) >= 3}
    return tokens


def _metadata_boost(
    metadata: dict[str, Any],
    query_tokens: set[str],
    *,
    priority_sheet_boost: float,
    token_match_boost: float,
) -> float:
    """Compute additive score boost based on metadata and query-token matches."""
    boost = 0.0

    if metadata.get("is_priority"):
        boost += priority_sheet_boost

    sheet_name = str(metadata.get("sheet_name", "")).lower()
    matched_intent = any(
        token.lower() in _SHEET_INTENT_TERMS.get(sheet_name, set())
        for token in query_tokens
    )
    if matched_intent:
        boost += token_match_boost / 2.0

    source_file = str(metadata.get("source_file", "")).upper()
    if source_file:
        file_tokens = _tokens_for_matching(source_file)
        if file_tokens.intersection(query_tokens):
            boost += token_match_boost

    return boost


@observe(name="rag-mda-templates-query")
def query_similar_templates(
    extraction_text: str,
    top_k: int = 3,
    collection_name: str = COLLECTION_NAME,
    chroma_path: str = CHROMA_PATH,
    use_query_augmentation: bool = True,
    query_augmentation_max_queries: int = 4,
    use_metadata_boost: bool = True,
    semantic_weight: float = 0.6,
    bm25_weight: float = 0.4,
    priority_sheet_boost: float = 0.15,
    token_match_boost: float = 0.2,
) -> list[str]:
    """Query ChromaDB for MDA templates similar to the given extraction text.

    Args:
        extraction_text: Text from a PDF extraction or user query to find
            similar templates for.
        top_k: Maximum number of similar documents to return.
        collection_name: ChromaDB collection name (defaults to ``mda_templates``).
        chroma_path: Path to the ChromaDB persistent storage directory.

    Returns:
        List of document strings from the most similar templates.

    Raises:
        ValueError: If *extraction_text* is empty.
        RuntimeError: If the collection is empty (seed first).
    """
    if not extraction_text.strip():
        msg = (
            "extraction_text must not be empty. "
            "Provide text from a PDF extraction or a search query."
        )
        raise ValueError(msg)

    scored = query_similar_templates_with_scores(
        extraction_text=extraction_text,
        top_k=top_k,
        collection_name=collection_name,
        chroma_path=chroma_path,
        use_query_augmentation=use_query_augmentation,
        query_augmentation_max_queries=query_augmentation_max_queries,
        use_metadata_boost=use_metadata_boost,
        semantic_weight=semantic_weight,
        bm25_weight=bm25_weight,
        priority_sheet_boost=priority_sheet_boost,
        token_match_boost=token_match_boost,
    )
    matched_docs = [item["content"] for item in scored]

    logger.info(
        "Query returned %d similar templates",
        len(matched_docs),
    )

    return matched_docs


@observe(name="rag-mda-templates-query-scored")
def query_similar_templates_with_scores(
    extraction_text: str,
    top_k: int = 3,
    collection_name: str = COLLECTION_NAME,
    chroma_path: str = CHROMA_PATH,
    use_query_augmentation: bool = True,
    query_augmentation_max_queries: int = 4,
    use_metadata_boost: bool = True,
    semantic_weight: float = 0.6,
    bm25_weight: float = 0.4,
    priority_sheet_boost: float = 0.15,
    token_match_boost: float = 0.2,
) -> list[dict[str, Any]]:
    """Query ChromaDB and return documents with distance scores and metadata.

    Same as :func:`query_similar_templates` but returns full result dicts
    needed by the RAG evaluator.

    Args:
        extraction_text: Text from a PDF extraction or user query.
        top_k: Maximum number of similar documents to return.
        collection_name: ChromaDB collection name.
        chroma_path: Path to the ChromaDB persistent storage directory.

    Returns:
        List of dicts with keys: ``content``, ``distance``, ``metadata``.

    Raises:
        ValueError: If *extraction_text* is empty.
        RuntimeError: If the collection is empty.
    """
    if not extraction_text.strip():
        raise ValueError("extraction_text must not be empty.")

    client = chromadb.PersistentClient(path=chroma_path)
    collection = client.get_or_create_collection(collection_name)

    doc_count = collection.count()
    if doc_count == 0:
        raise RuntimeError(
            f"ChromaDB collection '{collection_name}' is empty. "
            f"Run seed_mda_templates() first."
        )

    effective_k = min(top_k, doc_count)
    candidate_k = min(max(effective_k * 3, effective_k), doc_count)

    query_variants = [extraction_text]
    if use_query_augmentation:
        query_variants = _build_augmented_queries(
            extraction_text,
            max_queries=query_augmentation_max_queries,
        )

    logger.info(
        "Querying '%s' collection (top_k=%d, docs_available=%d, query_variants=%d)",
        collection_name,
        effective_k,
        doc_count,
        len(query_variants),
    )

    site_prefix = _extract_site_prefix(extraction_text)
    where_clause = {"prefix": site_prefix} if site_prefix else None
    semantic_rank_lists: list[list[str]] = []

    for query_variant in query_variants:
        semantic_results = collection.query(
            query_texts=[query_variant],
            n_results=candidate_k,
            where=where_clause,
            include=["metadatas", "distances"],
        )
        semantic_ids = semantic_results.get("ids", [[]])[0]
        semantic_rank_lists.append(semantic_ids)

    semantic_ids_fused = _reciprocal_rank_fusion(
        semantic_ids=[doc_id for rank_list in semantic_rank_lists for doc_id in rank_list],
        bm25_ids=[],
        semantic_weight=semantic_weight,
        bm25_weight=0.0,
    )

    _build_bm25_index(collection)
    bm25_ids: list[str] = []
    if _bm25_index is not None:
        bm25_rank_scores: dict[str, float] = {}
        for query_variant in query_variants:
            tokenized_query = _tokenize(query_variant)
            bm25_scores = _bm25_index.get_scores(tokenized_query)
            ranked_indices = sorted(
                range(len(bm25_scores)),
                key=lambda i: bm25_scores[i],
                reverse=True,
            )[:candidate_k]
            for rank, idx in enumerate(ranked_indices):
                doc_id = _bm25_doc_ids[idx]
                bm25_rank_scores[doc_id] = bm25_rank_scores.get(doc_id, 0.0) + (1.0 / (60 + rank + 1))

        bm25_ids = sorted(
            bm25_rank_scores.keys(),
            key=lambda doc_id: bm25_rank_scores[doc_id],
            reverse=True,
        )[:candidate_k]

    fused_ids = _reciprocal_rank_fusion(
        semantic_ids=semantic_ids_fused,
        bm25_ids=bm25_ids,
        semantic_weight=semantic_weight,
        bm25_weight=bm25_weight,
    )

    if not fused_ids:
        return []

    candidate_payload = _collection_get(
        collection,
        ids=fused_ids[:candidate_k],
        include=["documents", "metadatas"],
    )
    candidate_docs = candidate_payload.get("documents", [])
    candidate_metas = candidate_payload.get("metadatas", [])

    query_tokens = _tokens_for_matching(extraction_text)
    scored_candidates: list[tuple[float, dict[str, Any]]] = []
    for rank, (doc, meta) in enumerate(zip(candidate_docs, candidate_metas), start=1):
        metadata = dict(meta) if meta else {}
        base_score = 1.0 / (60 + rank)
        metadata_bonus = 0.0
        if use_metadata_boost:
            metadata_bonus = _metadata_boost(
                metadata,
                query_tokens,
                priority_sheet_boost=priority_sheet_boost,
                token_match_boost=token_match_boost,
            )
        total_score = base_score + metadata_bonus
        scored_candidates.append((
            total_score,
            {
                "content": doc,
                "distance": 0.0,
                "metadata": metadata,
                "score": total_score,
                "base_score": base_score,
                "metadata_bonus": metadata_bonus,
            },
        ))

    scored_candidates.sort(key=lambda item: item[0], reverse=True)
    output = [item[1] for item in scored_candidates[:effective_k]]

    logger.info(
        "Scored query returned %d results (query_variants=%d)",
        len(output),
        len(query_variants),
    )

    return output
