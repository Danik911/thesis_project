"""XLSX exporter for MDA (Method Definition and Analysis) templates.

Exports an MDATemplate to a 4-sheet LabWare MDA format workbook using openpyxl.
Sheets: Analysis, Component, Calc Variable, Calculation.

Styling matches LabWare conventions:
  - Dark green header row with white bold text
  - Thin borders on all cells
  - Auto-width columns (capped at 50 characters)

GAMP-5 Category 5: Custom pharmaceutical software component.
No fallback logic -- all errors propagate with full diagnostics.
"""

from __future__ import annotations

import io
import logging
from collections.abc import Sequence
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .mda_schema import MDATemplate

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(
    start_color="1F4E3D",
    end_color="1F4E3D",
    fill_type="solid",
)
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_MAX_COLUMN_WIDTH = 50

# Sheet definitions: (sheet_name, column_keys)
# Column keys correspond to fields returned by model_dump() on each item.
_SHEET_DEFINITIONS: list[tuple[str, list[str]]] = [
    (
        "Analysis",
        [
            "name",
            "version",
            "group_name",
            "active",
            "reported_name",
            "common_name",
            "analysis_type",
            "description",
        ],
    ),
    (
        "Component",
        [
            "analysis",
            "component_name",
            "version",
            "order_number",
            "result_type",
            "units",
            "minimum",
            "maximum",
            "uses_instrument",
            "instrument_group",
            "auto_calc",
            "list_key",
            "reportable",
            "optional",
        ],
    ),
    (
        "Calc Variable",
        [
            "analysis",
            "component",
            "name",
            "version",
            "reference_type",
            "reference_analysis",
            "reference_component",
            "return_value",
            "scope",
            "function",
        ],
    ),
    (
        "Calculation",
        [
            "analysis",
            "component",
            "version",
            "description",
            "source_code",
            "calculation_type",
        ],
    ),
]


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _format_cell_value(value: Any) -> str:
    """Convert a Python value to its XLSX string representation.

    Rules:
      - bool  -> "TRUE" / "FALSE"
      - list  -> comma-separated string
      - None  -> "" (empty string)
      - other -> str(value)

    Raises:
        TypeError: If the value cannot be converted to string.
    """
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    if value is None:
        return ""
    return str(value)


def _write_header_row(
    ws: Worksheet,
    columns: list[str],
) -> None:
    """Write styled header row to the worksheet.

    Args:
        ws: Target worksheet.
        columns: Column header names.
    """
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER


def _write_data_rows(
    ws: Worksheet,
    columns: list[str],
    items: Sequence[Any],
) -> None:
    """Write data rows from model instances to the worksheet.

    Each item is dumped via model_dump() and only the specified columns
    are extracted in order.

    Args:
        ws: Target worksheet.
        columns: Column keys matching model_dump() field names.
        items: Sequence of Pydantic model instances with model_dump().
    """
    for row_idx, item in enumerate(items, start=2):
        data = item.model_dump()
        for col_idx, col_key in enumerate(columns, start=1):
            raw_value = data.get(col_key)
            cell_value = _format_cell_value(raw_value)
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = _THIN_BORDER


def _auto_fit_columns(ws: Worksheet, columns: list[str]) -> None:
    """Auto-fit column widths based on content, capped at _MAX_COLUMN_WIDTH.

    Scans header and all data rows to determine the maximum content length
    per column, then sets the column width accordingly.

    Args:
        ws: Target worksheet.
        columns: Column header names (used for initial width calculation).
    """
    for col_idx in range(1, len(columns) + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(
            min_col=col_idx,
            max_col=col_idx,
            min_row=1,
            max_row=ws.max_row,
        ):
            for cell in row:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)

        # Add small padding (2 chars) and cap at maximum
        adjusted_width = min(max_length + 2, _MAX_COLUMN_WIDTH)
        ws.column_dimensions[col_letter].width = adjusted_width


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def export_mda_to_xlsx(mda: MDATemplate) -> bytes:
    """Export an MDATemplate to a 4-sheet LabWare MDA format XLSX workbook.

    Creates an in-memory XLSX workbook with the following sheets:
      1. Analysis    -- top-level test method definitions
      2. Component   -- individual test parameters
      3. Calc Variable -- variables for LIMS Basic calculations
      4. Calculation -- LIMS Basic calculation code

    Args:
        mda: A validated MDATemplate instance containing analyses,
             components, calc_variables, and calculations.

    Returns:
        Raw bytes of the XLSX file, ready for streaming or saving.

    Raises:
        openpyxl.utils.exceptions.InvalidFileException: If workbook
            creation fails.
        AttributeError: If mda items lack model_dump() method.
    """
    logger.info(
        "Exporting MDA template to XLSX: %d analyses, %d components, "
        "%d calc_variables, %d calculations",
        len(mda.analyses),
        len(mda.components),
        len(mda.calc_variables),
        len(mda.calculations),
    )

    # Map sheet names to their data source on the MDATemplate
    data_sources: dict[str, Sequence[Any]] = {
        "Analysis": mda.analyses,
        "Component": mda.components,
        "Calc Variable": mda.calc_variables,
        "Calculation": mda.calculations,
    }

    wb = openpyxl.Workbook()

    for sheet_idx, (sheet_name, columns) in enumerate(_SHEET_DEFINITIONS):
        # First sheet reuses the default active sheet; subsequent sheets
        # are created explicitly.
        if sheet_idx == 0:
            ws = wb.active
            if ws is None:
                msg = (
                    "Workbook has no active sheet. "
                    "openpyxl Workbook() should create one by default."
                )
                raise RuntimeError(msg)
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        items = data_sources[sheet_name]

        _write_header_row(ws, columns)
        _write_data_rows(ws, columns, items)
        _auto_fit_columns(ws, columns)

        logger.debug(
            "Sheet '%s': wrote %d data rows with %d columns",
            sheet_name,
            len(items),
            len(columns),
        )

    # Serialize to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    xlsx_bytes = buffer.read()

    logger.info(
        "XLSX export complete: %d bytes",
        len(xlsx_bytes),
    )

    return xlsx_bytes
