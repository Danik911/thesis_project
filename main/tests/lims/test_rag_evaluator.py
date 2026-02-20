"""Tests for main.src.lims.rag_evaluator -- RAG evaluation framework."""

from __future__ import annotations

import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from main.src.lims.rag_evaluator import (
    EvalQuery,
    EvalResult,
    EvalSummary,
    RAGEvaluator,
)


class TestEvalMetrics:
    """Test metric calculations without ChromaDB."""

    def test_hit_rate_all_hits(self):
        results = [
            EvalResult(
                query=EvalQuery("q1", "file1.xlsx"),
                hit=True, reciprocal_rank=1.0, precision_at_k=1.0,
                distances=[0.1], retrieved_sources=["file1.xlsx"],
            ),
            EvalResult(
                query=EvalQuery("q2", "file2.xlsx"),
                hit=True, reciprocal_rank=0.5, precision_at_k=0.5,
                distances=[0.2, 0.3], retrieved_sources=["file2.xlsx", "other.xlsx"],
            ),
        ]
        total = len(results)
        hit_rate = sum(1 for r in results if r.hit) / total
        mrr = sum(r.reciprocal_rank for r in results) / total
        assert hit_rate == 1.0
        assert mrr == 0.75

    def test_hit_rate_no_hits(self):
        results = [
            EvalResult(
                query=EvalQuery("q1", "file1.xlsx"),
                hit=False, reciprocal_rank=0.0, precision_at_k=0.0,
                distances=[0.9], retrieved_sources=["other.xlsx"],
            ),
        ]
        hit_rate = sum(1 for r in results if r.hit) / len(results)
        assert hit_rate == 0.0

    def test_precision_at_k(self):
        # 2 out of 3 results match expected source
        result = EvalResult(
            query=EvalQuery("q", "target.xlsx"),
            hit=True,
            reciprocal_rank=1.0,
            precision_at_k=2 / 3,
            distances=[0.1, 0.2, 0.5],
            retrieved_sources=["target.xlsx", "target.xlsx", "other.xlsx"],
        )
        assert abs(result.precision_at_k - 0.6667) < 0.001

    def test_mrr_calculation(self):
        # Expected source at rank 3
        result = EvalResult(
            query=EvalQuery("q", "target.xlsx"),
            hit=True,
            reciprocal_rank=1 / 3,
            precision_at_k=1 / 3,
            distances=[0.1, 0.2, 0.3],
            retrieved_sources=["a.xlsx", "b.xlsx", "target.xlsx"],
        )
        assert abs(result.reciprocal_rank - 0.3333) < 0.001


class TestEvalSummaryStr:
    def test_str_repr(self):
        summary = EvalSummary(
            total_queries=10,
            hit_rate=0.8,
            mrr=0.65,
            mean_precision_at_k=0.5,
            mean_distance=0.3456,
            top_k=3,
        )
        s = str(summary)
        assert "top_k=3" in s
        assert "queries=10" in s
        assert "hit_rate=0.800" in s
        assert "mrr=0.650" in s


class TestBuildMdaEvalQueries:
    """Test query building from demo_data directory."""

    def test_builds_queries_from_matching_pairs(self, tmp_path):
        # Create matching PDF + XLSX pairs
        (tmp_path / "AND_TEST.pdf").touch()
        (tmp_path / "AND_TEST.xlsx").touch()
        wb = openpyxl.Workbook()
        wb.save(str(tmp_path / "AND_TEST.xlsx"))
        wb.close()

        (tmp_path / "FRE_OTHER.pdf").touch()
        (tmp_path / "FRE_OTHER.xlsx").touch()
        wb2 = openpyxl.Workbook()
        wb2.save(str(tmp_path / "FRE_OTHER.xlsx"))
        wb2.close()

        # XLSX without matching PDF should be excluded
        (tmp_path / "ORPHAN.xlsx").touch()
        wb3 = openpyxl.Workbook()
        wb3.save(str(tmp_path / "ORPHAN.xlsx"))
        wb3.close()

        # Mock ChromaDB to avoid needing a real collection
        with patch("main.src.lims.rag_evaluator.chromadb") as mock_chroma:
            mock_collection = MagicMock()
            mock_collection.count.return_value = 10
            mock_client = MagicMock()
            mock_client.get_or_create_collection.return_value = mock_collection
            mock_chroma.PersistentClient.return_value = mock_client

            evaluator = RAGEvaluator(
                chroma_path=str(tmp_path / "chroma"),
                collection_name="test",
            )
            queries = evaluator.build_mda_eval_queries(str(tmp_path))

        assert len(queries) == 2
        expected_files = {q.expected_source_file for q in queries}
        assert "AND_TEST.xlsx" in expected_files
        assert "FRE_OTHER.xlsx" in expected_files
        assert "ORPHAN.xlsx" not in expected_files

    def test_query_text_format(self, tmp_path):
        (tmp_path / "AND_ACS_DYE-LAB-2499.pdf").touch()
        (tmp_path / "AND_ACS_DYE-LAB-2499.xlsx").touch()
        wb = openpyxl.Workbook()
        wb.save(str(tmp_path / "AND_ACS_DYE-LAB-2499.xlsx"))
        wb.close()

        with patch("main.src.lims.rag_evaluator.chromadb") as mock_chroma:
            mock_collection = MagicMock()
            mock_collection.count.return_value = 10
            mock_client = MagicMock()
            mock_client.get_or_create_collection.return_value = mock_collection
            mock_chroma.PersistentClient.return_value = mock_client

            evaluator = RAGEvaluator(
                chroma_path=str(tmp_path / "chroma"),
                collection_name="test",
            )
            queries = evaluator.build_mda_eval_queries(str(tmp_path))

        assert len(queries) == 1
        # Underscores and hyphens should be replaced with spaces
        assert queries[0].query_text == "AND ACS DYE LAB 2499"

    def test_file_not_found(self):
        with patch("main.src.lims.rag_evaluator.chromadb") as mock_chroma:
            mock_collection = MagicMock()
            mock_collection.count.return_value = 10
            mock_client = MagicMock()
            mock_client.get_or_create_collection.return_value = mock_collection
            mock_chroma.PersistentClient.return_value = mock_client

            evaluator = RAGEvaluator(
                chroma_path="/tmp/fake_chroma",
                collection_name="test",
            )
            with pytest.raises(FileNotFoundError):
                evaluator.build_mda_eval_queries("/nonexistent/dir")


class TestEvaluateWithMockCollection:
    """Test evaluate() with mocked ChromaDB collection."""

    def test_evaluate_perfect_retrieval(self):
        with patch("main.src.lims.rag_evaluator.chromadb") as mock_chroma:
            mock_collection = MagicMock()
            mock_collection.count.return_value = 10
            mock_collection.query.return_value = {
                "metadatas": [[{"source_file": "target.xlsx"}]],
                "distances": [[0.15]],
            }
            mock_client = MagicMock()
            mock_client.get_or_create_collection.return_value = mock_collection
            mock_chroma.PersistentClient.return_value = mock_client

            evaluator = RAGEvaluator(chroma_path="/tmp/test", collection_name="test")
            queries = [EvalQuery("test query", "target.xlsx")]
            summary = evaluator.evaluate(queries, top_k=1)

        assert summary.hit_rate == 1.0
        assert summary.mrr == 1.0
        assert summary.mean_precision_at_k == 1.0

    def test_evaluate_miss(self):
        with patch("main.src.lims.rag_evaluator.chromadb") as mock_chroma:
            mock_collection = MagicMock()
            mock_collection.count.return_value = 10
            mock_collection.query.return_value = {
                "metadatas": [[{"source_file": "wrong.xlsx"}]],
                "distances": [[0.9]],
            }
            mock_client = MagicMock()
            mock_client.get_or_create_collection.return_value = mock_collection
            mock_chroma.PersistentClient.return_value = mock_client

            evaluator = RAGEvaluator(chroma_path="/tmp/test", collection_name="test")
            queries = [EvalQuery("test query", "target.xlsx")]
            summary = evaluator.evaluate(queries, top_k=1)

        assert summary.hit_rate == 0.0
        assert summary.mrr == 0.0

    def test_empty_queries_raises(self):
        with patch("main.src.lims.rag_evaluator.chromadb") as mock_chroma:
            mock_collection = MagicMock()
            mock_collection.count.return_value = 10
            mock_client = MagicMock()
            mock_client.get_or_create_collection.return_value = mock_collection
            mock_chroma.PersistentClient.return_value = mock_client

            evaluator = RAGEvaluator(chroma_path="/tmp/test", collection_name="test")
            with pytest.raises(ValueError, match="queries list must not be empty"):
                evaluator.evaluate([], top_k=3)
